#!/usr/bin/env python3
"""트래지디 루퍼 각본가 봇 — 스캐폴딩 번들 생성기

입력: 포크 레포의 data/*.jsonc + 정발 용어 대조표 xlsx
출력: /mnt/user-data/outputs/tragedy-looper-mastermind/
"""
import re, json, glob, os, shutil
from collections import Counter, OrderedDict
from openpyxl import load_workbook

DATA = "./tragedy-looper-ko/data"   # 포크 레포 클론 경로
XLSX = "./tragedy-looper-정발용어-대조표-v4.xlsx"
OUT  = "."

L = lambda f: json.loads(re.sub(r'^\s*//.*$', '', open(f).read(), flags=re.M))

# ─────────────────────────────────────────────────────────── 레포 데이터 수집
def collect(kind, plural):
    out = OrderedDict()
    for f in sorted(glob.glob(f"{DATA}/*/{kind}.jsonc")):
        st = os.path.basename(os.path.dirname(f))
        try:
            for o in L(f)[plural]:
                o["_set"] = st
                out[o["id"]] = o
        except Exception:
            pass
    return out

chars     = collect("characters", "characters")
roles     = collect("roles", "roles")
plots     = collect("plots", "plots")
incidents = collect("incidents", "incidents")
tragedys  = collect("tragedys", "tragedys")

# ─────────────────────────────────────────────────────────── 정발 용어 흡수
wb = load_workbook(XLSX, read_only=True)
def sheet(n):
    ws = wb[n]; it = ws.iter_rows(values_only=True); hdr = next(it)
    return [dict(zip(hdr, r)) for r in it if any(v is not None for v in r)]

KEY = "영문 키 (수정 금지)"
def ko_map(sheet_name, repo_dict):
    """영문 name → 한국어. repo id로 재키잉하고 미매칭을 보고."""
    by_name = {o["name"]: i for i, o in repo_dict.items()}
    hit, miss = {}, []
    for r in sheet(sheet_name):
        k = (r.get(KEY) or "").strip()
        ko = (r.get("정발 용어") or "").strip()
        grade = (r.get("등급") or "").strip()
        if not k:
            continue
        rid = by_name.get(k)
        if rid is None:
            if ko: miss.append(k)
            continue
        if ko:
            hit[rid] = {"ko": ko, "grade": grade or "미지정"}
    return hit, miss

ko_char, miss_char = ko_map("캐릭터", chars)
ko_role, miss_role = ko_map("역할", roles)
ko_inc,  miss_inc  = ko_map("사건", incidents)
ko_plot, miss_plot = ko_map("플롯", plots)
ko_trag, miss_trag = ko_map("참극 세트", tragedys)

ko_misc = {}
for sn in ("핵심 개념", "키워드·태그", "게임 용어·표기"):
    for r in sheet(sn):
        k = (r.get(KEY) or r.get("영문 키") or "")
        k = k.strip() if isinstance(k, str) else ""
        ko = (r.get("정발 용어") or "").strip()
        if k and ko:
            ko_misc[k] = {"ko": ko, "grade": (r.get("등급") or "미지정").strip()}

# ─────────────────────────────────────────────────────────── 한국어 발매 구성
# 근거: 주인공 설명서 16p「캐릭터 카드 (24장)」, 41p 프로모 2종 명시,
#       기본/미스터리서클/미드나이트존 추리 참조표 캐릭터 목록 대조.
KO_BASE_BOX = [
    "boyStudent","girlStudent","richStudent","classRep","mysteryBoy",
    "shrineMaiden","alien","godlyBeing","policeOfficer","officeWorker",
    "informer","popIdol","journalist","boss","doctor","patient","nurse",
    "henchman","teacher","transferStudent","soldier","blackCat",
    "forensicSpecialist","ai",
]
KO_PROMO = ["scientist", "illusion"]
KO_TRAGEDY_SETS = {
    "firstSteps":     "본판",
    "basicTragedy":   "본판",
    "mysteryCircle":  "미드나이트 서클",
    "midnightZone":   "미드나이트 서클",
    "hauntedStage":   "헌티드 스테이지",
    "cosmicEvil":     "위어드 미솔로지",
}

release = {
    "_note": "레포의 edition/폴더 구분은 원작 발매 단위이며 한국어 실물 박스와 다름. "
             "이 파일이 한국어 기준 축이다. 24장 구성은 주인공 설명서 16p 및 "
             "추리 참조표 대조로 도출 — 실물 검수 권장.",
    "characters": {
        "본판": KO_BASE_BOX,
        "프로모": KO_PROMO,
        "미발매": sorted(set(chars) - set(KO_BASE_BOX) - set(KO_PROMO)),
    },
    "tragedySets": KO_TRAGEDY_SETS,
}

# ─────────────────────────────────────────────────────────── 타이밍 → 단계
PHASE = {
    "Day Start": "P1_ROUND_START", "First Day": "P1_ROUND_START",
    "Mastermind Action step": "P2_MASTERMIND_ACTION",
    "Card resolve": "P4_RESOLVE",
    "Mastermind Ability": "P5_MASTERMIND_ABILITY",
    "Goodwill ablility step": "P6_GOODWILL",
    "After Goodwill Ability used": "P6_GOODWILL",
    "Incident step": "P7_INCIDENT", "Incident trigger": "P7_INCIDENT",
    "Day End": "P9_ROUND_END",
    # 단계 밖 훅
    "Always": "ALWAYS", "Loop Start": "LOOP_START", "Loop End": "LOOP_END",
    "Last Day": "LAST_DAY", "On character death": "ON_DEATH",
    "When this role is to be reveald": "ON_REVEAL",
    "Final Guess": "FINAL_GUESS", "Script creation": "SCRIPT_BUILD",
}
KIND = {
    "Mandatory": "mandatory", "Optional": "optional",
    "Loss condition: Tragedy": "lossTragedy",
    "Optional Loss condition: Tragedy": "lossTragedy",
    "Optional Loss condition: Protagonists Death": "lossDeath",
    "Mandatory Loss condition: Protagonists Death": "lossDeath",
    "Delayed Loss condition: Protagonists Death": "lossDeath",
    "Script creation": "scriptBuild",
}

def esc(s):
    return (s or "").replace("\\", "\\\\").replace("`", "\\`").replace("${", "\\${")

def hooks_of(obj, field):
    out = []
    for a in obj.get(field, []):
        timings = a.get("timing") or ["Always"]
        for t in timings:
            out.append({
                "phase": PHASE.get(t, "UNKNOWN"),
                "kind": KIND.get(a.get("type", ""), "mandatory"),
                "timing_raw": t,
                "type_raw": a.get("type", ""),
                "pre": a.get("prerequisite", ""),
                "desc": a.get("description", ""),
                "timesPerLoop": a.get("timesPerLoop"),
            })
    return out

# ─────────────────────────────────────────────────────────── 기본편 대상 선정
bt = tragedys["basicTragedy"]
bt_plots = list(dict.fromkeys(bt["mainPlots"] + bt["subPlots"]))
bt_roles = ["person"]
for p in bt_plots:
    for r in plots[p].get("roles", {}):
        if r not in bt_roles:
            bt_roles.append(r)
bt_incidents = list(bt["incidents"])

def kname(d, i):
    return d.get(i, {}).get("ko") or f"({i})"

# ─────────────────────────────────────────────────────────── 코드 생성
def gen_impl(title, ids, src, field, kind_word):
    lines = [
        "// ⚠️ 자동 생성 스캐폴딩 — 구조는 생성기가, 로직은 사람이.",
        "//    재생성해도 when/effect 는 덮어쓰지 않도록 주의할 것.",
        "//    source 는 원본 영문 텍스트(수정 금지). ko 는 정발 용어.",
        "",
        'import type { GameState, CharacterId, Hook } from "../types";',
        "",
        f"/** {title} — 총 {len(ids)}건 */",
    ]
    total_hooks = 0
    body = []
    for i in ids:
        o = src[i]
        hs = hooks_of(o, field)
        total_hooks += len(hs)
        ko = kname({"characters": ko_char, "roles": ko_role,
                    "incidents": ko_inc, "plots": ko_plot}[kind_word], i)
        meta = []
        if o.get("goodwillRefusel"): meta.append(f'goodwillRefusal: "{o["goodwillRefusel"]}"')
        if o.get("max") is not None: meta.append(f'max: {o["max"]}')
        if o.get("tags"): meta.append(f'tags: {json.dumps(o["tags"])}')
        if o.get("roles"): meta.append(f'addsRoles: {json.dumps(o["roles"])}')
        body.append(f'  // ── {ko} ({o["name"]})')
        body.append(f'  {i}: {{')
        body.append(f'    ko: "{ko}",')
        for m in meta:
            body.append(f"    {m},")
        if not hs:
            body.append("    hooks: [], // 능력 없음")
        else:
            body.append("    hooks: [")
            for h in hs:
                body.append("      {")
                body.append(f'        phase: "{h["phase"]}",')
                body.append(f'        kind: "{h["kind"]}",')
                if h["timesPerLoop"]:
                    body.append(f'        timesPerLoop: {h["timesPerLoop"]},')
                body.append("        source: {")
                body.append(f'          timing: "{esc(h["timing_raw"])}",')
                if h["pre"]:
                    body.append(f'          prerequisite: `{esc(h["pre"])}`,')
                if h["desc"]:
                    body.append(f'          description: `{esc(h["desc"])}`,')
                body.append("        },")
                body.append("        // TODO(구현): 위 source 를 술어/효과로 옮길 것")
                body.append("        when: (_s: GameState, _self: CharacterId) => false,")
                body.append("        effect: (_s: GameState, _self: CharacterId) => { throw new Error('unimplemented'); },")
                body.append("      },")
            body.append("    ],")
        body.append("  },")
    varname = {"roles": "ROLE_IMPL", "plots": "PLOT_IMPL",
               "incidents": "INCIDENT_IMPL"}[kind_word]
    lines.append(f"export const {varname}: Record<string, {{")
    lines.append("  ko: string;")
    lines.append("  goodwillRefusal?: 'Optional' | 'Mandatory';")
    lines.append("  max?: number;")
    lines.append("  tags?: string[];")
    lines.append("  addsRoles?: Record<string, number>;")
    lines.append("  hooks: Hook[];")
    lines.append("}> = {")
    lines += body
    lines.append("};")
    lines.append("")
    return "\n".join(lines), total_hooks

os.makedirs(f"{OUT}/src/impl", exist_ok=True)
os.makedirs(f"{OUT}/src/engine", exist_ok=True)
os.makedirs(f"{OUT}/data", exist_ok=True)

r_src, r_n = gen_impl("기본편 역할", bt_roles, roles, "abilities", "roles")
p_src, p_n = gen_impl("기본편 룰(플롯)", bt_plots, plots, "rules", "plots")
i_src, i_n = gen_impl("기본편 사건", bt_incidents, incidents, "effect", "incidents")
open(f"{OUT}/src/impl/roles.ts", "w").write(r_src)
open(f"{OUT}/src/impl/plots.ts", "w").write(p_src)
open(f"{OUT}/src/impl/incidents.ts", "w").write(i_src)

# ─────────────────────────────────────────────────────────── 정적 데이터 JSON
def char_static(i):
    c = chars[i]
    return {
        "id": i, "en": c["name"], "ko": ko_char.get(i, {}).get("ko"),
        "paranoiaLimit": c.get("paranoiaLimit"),
        "startLocation": c.get("startLocation"),
        "forbiddenLocation": c.get("forbiddenLocation", []),
        "tags": c.get("tags", []),
        "plotLessRole": bool(c.get("plotLessRole")),
        "comesInLater": bool(c.get("comesInLater")),
        "scriptSpecified": c.get("scriptSpecified"),
        "goodwillAbilities": [
            {"rank": a.get("goodwillRank"), "en": a.get("description"),
             "timesPerLoop": a.get("timesPerLoop"),
             "restrictedToLocation": a.get("restrictedToLocation")}
            for a in c.get("abilities", [])
        ],
    }

json.dump({i: char_static(i) for i in chars}, open(f"{OUT}/data/characters.json", "w"),
          ensure_ascii=False, indent=2)
json.dump(release, open(f"{OUT}/data/ko-release.json", "w"), ensure_ascii=False, indent=2)
json.dump({"characters": ko_char, "roles": ko_role, "incidents": ko_inc,
           "plots": ko_plot, "tragedySets": ko_trag, "misc": ko_misc},
          open(f"{OUT}/data/ko-terms.json", "w"), ensure_ascii=False, indent=2)

# 참극 세트별 시나리오 (테스트 픽스처)
base_game_scripts = L(f"{DATA}/base-game/scripts.jsonc")["scripts"]
fs_scripts = [s for s in base_game_scripts
              if s.get("tragedySet") == "firstSteps"]
bt_scripts = [s for s in base_game_scripts
              if s.get("tragedySet") == "basicTragedy"]
json.dump(fs_scripts, open(f"{OUT}/data/first-steps-scripts.json", "w"),
          ensure_ascii=False, indent=2)
json.dump(bt_scripts, open(f"{OUT}/data/basic-tragedy-scripts.json", "w"),
          ensure_ascii=False, indent=2)

# ─────────────────────────────────────────────────────────── 리포트
report = {
    "roles": {"count": len(bt_roles), "hooks": r_n},
    "plots": {"count": len(bt_plots), "hooks": p_n},
    "incidents": {"count": len(bt_incidents), "hooks": i_n},
    "scripts": len(bt_scripts),
    "first_steps_scripts": len(fs_scripts),
    "characters_total": len(chars),
    "ko_missing_characters": sorted(i for i in chars if i not in ko_char),
    "xlsx_unmatched": {"characters": miss_char, "roles": miss_role,
                       "incidents": miss_inc, "plots": miss_plot},
}
json.dump(report, open("/home/claude/report.json", "w"), ensure_ascii=False, indent=2)
print(json.dumps(report, ensure_ascii=False, indent=2))
